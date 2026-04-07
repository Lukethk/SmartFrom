from io import BytesIO

from openpyxl import Workbook, load_workbook

from app.models.entities import Cell, Document, Extraction, MappingProfile, Template
from app.services.exporter import export_batch_to_xlsx


def test_export_injects_values() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["B1"] = "first_name"
    ws["C1"] = "last_name"
    buffer = BytesIO()
    wb.save(buffer)

    template = Template(name="T", filename="t.xlsx", owner_id="u", workbook_blob=buffer.getvalue())
    profile = MappingProfile(template_id="t", name="map", sheet_name="Sheet1", start_row=2, mapping_json={"first_name": "B", "last_name": "C"})
    doc = Document(batch_id="b", filename="f", mime_type="image/png")
    extraction = Extraction(document_id="d", template_id="t")
    cells = [
        Cell(extraction_id="e", logical_key="first_name", raw_text="Juan", normalized_value="Juan", confidence=0.9, source_bbox={}, is_validated=True),
        Cell(extraction_id="e", logical_key="last_name", raw_text="Perez", normalized_value="Perez", confidence=0.9, source_bbox={}, is_validated=True),
    ]

    out = export_batch_to_xlsx(template, profile, [(doc, extraction, cells)], include_only_validated=True)
    wb_out = load_workbook(BytesIO(out))
    assert wb_out["Sheet1"]["B2"].value == "Juan"
    assert wb_out["Sheet1"]["C2"].value == "Perez"
