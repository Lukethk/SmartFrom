from io import BytesIO

from openpyxl import load_workbook

from app.models.entities import Cell, Document, Extraction, MappingProfile, Template


def export_batch_to_xlsx(
    template: Template,
    mapping_profile: MappingProfile,
    extraction_rows: list[tuple[Document, Extraction, list[Cell]]],
    include_only_validated: bool,
) -> bytes:
    wb = load_workbook(BytesIO(template.workbook_blob))
    ws = wb[mapping_profile.sheet_name] if mapping_profile.sheet_name in wb.sheetnames else wb.active

    row = mapping_profile.start_row
    for _, _, cells in extraction_rows:
        values = {}
        for cell in cells:
            if include_only_validated and not cell.is_validated:
                continue
            values[cell.logical_key] = cell.normalized_value or cell.raw_text

        for logical_key, column_ref in mapping_profile.mapping_json.items():
            val = values.get(logical_key, "")
            ws[f"{column_ref}{row}"] = val
        row += 1

    out = BytesIO()
    wb.save(out)
    return out.getvalue()
